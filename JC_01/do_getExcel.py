# -*- coding: utf-8 -*-
'''
@File  : do_getExcel.py
@Author: 汪先锦
@Date  : 2019-12-14 19:13
@Software: PyCharm
@Desc  :
'''
from JC_01 import filePath
from openpyxl import load_workbook
class DoExcel:
    def get_data(self,file_name,sheet_name):
        wb=load_workbook(file_name)
        sheet=wb[sheet_name]

        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data={}
            row_data['API']=sheet.cell(i,1).value
            row_data['secretkey']=sheet.cell(i,2).value
            row_data['body']=sheet.cell(i,3).value

            test_data.append(row_data)
            print(row_data)
        return test_data
if __name__ == '__main__':
    d=DoExcel()
    d.get_data(filePath.filePath('JC_01','sign.xlsx'),'Sheet2')